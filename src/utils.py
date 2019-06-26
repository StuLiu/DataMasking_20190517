'''
--------------------------------------------------------
@File    :   masker.py    
@Contact :   1183862787@qq.com
@License :   (C)Copyright 2017-2018, CS, WHU

@Modify Time : 2019/5/18 23:51     
@Author      : Liu Wang    
@Version     : 1.0   
@Desciption  : None
--------------------------------------------------------  
'''

from gmssl.sm4 import CryptSM4, SM4_ENCRYPT, SM4_DECRYPT
import pickle
from openpyxl import load_workbook, Workbook
import time

def _encrypt(key_str, data_str): # ->bytes
	if data_str == None:
		return None
	assert len(key_str) == 16
	key_bytes = bytes(key_str, encoding='utf-8')
	data_bytes = bytes(data_str, encoding='utf-8')
	crypt_sm4 = CryptSM4()
	crypt_sm4.set_key(key_bytes, SM4_ENCRYPT)
	return crypt_sm4.crypt_ecb(data_bytes)  # bytes类型

def _decrypt(key_str, data_bytes):   # ->str
	if data_bytes == None:
		return None
	assert len(key_str) == 16
	key_bytes = bytes(key_str, encoding='utf-8')
	crypt_sm4 = CryptSM4()
	crypt_sm4.set_key(key_bytes, SM4_DECRYPT)
	decrypted_data = crypt_sm4.crypt_ecb(data_bytes)  # bytes类型
	return str(decrypted_data, encoding='utf-8')

def append_zero(key_str)->str:
	"""对不够16位的密码做补零处理"""
	if len(key_str) < 16:
		key_str += '0' * (16 - len(key_str))
	return key_str

def save_pickle(file_name, dict_obj):
	assert type(dict_obj) == type(dict())
	with open(file_name, "wb") as file:
		pickle.dump(dict_obj, file)
	pass

def load_pickle(file_name)->dict:
	with open(file_name, "rb") as file:
		data = pickle.load(file)
		assert type(data) == type(dict())
		return data

def mask_row(key_str, year, row)->(list, dict):
	hash_bytes = {}     # the dict with key:hashcode and value:bytes
	if year == '17' or year == '18':
		indexs = [7, 20, 22, 24, 26, 33, 34, 39, 48, 53]    # 必须加密字段的下标
	else:
		indexs = []
	# 处理条件加密字段
	if year == '17' or year == '18':
		# 如果抽样环节为生产则加密
		if row[36] == '生产':
			encrypt_bytes = _encrypt(key_str, str(row[0]))
			encrypt_hash = str(hash(encrypt_bytes))
			hash_bytes[encrypt_hash] = encrypt_bytes
			row[0] = encrypt_hash
			encrypt_bytes = _encrypt(key_str, str(row[6]))
			encrypt_hash = str(hash(encrypt_bytes))
			hash_bytes[encrypt_hash] = encrypt_bytes
			row[6] = encrypt_hash
		# 加密样品名称中的商标
		row[8] = str(row[8]).replace(str(row[39]), '********')
	# 处理必须加密字段
	for index in indexs:
		encrypt_bytes = _encrypt(key_str, str(row[index]))
		encrypt_hash = str(hash(encrypt_bytes))
		row[index] = encrypt_hash
		hash_bytes[encrypt_hash] = encrypt_bytes
	return row, hash_bytes

def demask_row(key_str, year, row, hash_bytes)->list:
	"""data_rows:list()"""
	# print(data_rows)
	if year == '17' or year == '18':
		indexs = [7, 20, 22, 24, 26, 33, 34, 39, 48, 53]    # 必须解密字段的下标
	else:
		indexs = []
	# 处理必须加密字段
	for index in indexs:
		# print(hash_bytes[row[index]])
		row[index] = _decrypt(key_str, hash_bytes[str(row[index])])
	# 处理条件加密字段
	if year == '17' or year == '18':
		# 如果是生产环节解密
		if row[36] == '生产':
			row[0] = _decrypt(key_str, hash_bytes[str(row[0])])
			row[6] = _decrypt(key_str, hash_bytes[str(row[6])])
		# 使用解密后的商标替换样品名称中的屏蔽字段
		row[8] = str(row[8]).replace('********', str(row[39]))
	return row

# def read_xlsx(file_path, have_head=True)->dict:
# 	result = {}
# 	wb = load_workbook(file_path)
# 	# print('wb.sheetnames:',wb.sheetnames)
# 	for sheetname in wb.sheetnames:
# 		sheet = wb[sheetname]
# 		rows_list = []
# 		for row in sheet.values:
# 			rows_list.append(list(row))
# 		result[sheetname] = rows_list
# 	return result
#
# def save_xlsx(file_path, data_dict)->None:
# 	wb = Workbook()
# 	for key, values in data_dict.items():
# 		ws = wb.create_sheet(title=key)
# 		for row in values:
# 			ws.append(row)
# 	wb.save(file_path)

# def do_masking():
# 	wb_read = load_workbook('17.18_bigdata.xlsx', read_only=True)
# 	wb_write = Workbook(write_only=True)
# 	hash_bytes = load_pickle('mapping.pkl')
# 	# get the count of all data rows
# 	row_count = 0
# 	current_count = 1
# 	for sheetname in wb_read.sheetnames:
# 		sheet_read = wb_read[sheetname]
# 		row_count += sheet_read.max_row
# 	# read data and do masking, and then save the masked rows
# 	for sheetname in wb_read.sheetnames:
# 		print('processing sheet {}:'.format(sheetname))
# 		sheet_read = wb_read[sheetname]
# 		# sheet_row_count = sheet_read.max_row
# 		# print(sheet_row_count)
# 		sheet_write = wb_write.create_sheet(title=sheetname)
# 		rows_read = sheet_read.rows
# 		for row in rows_read:
# 			row_values = []
# 			for cell in row:
# 				row_values.append(cell.value)
# 			# do masking
# 			if current_count > 1:
# 				masked_row, hash_bytes_added = mask_row('1234567890123456', sheetname, row_values)
# 				hash_bytes.update(hash_bytes_added)
# 				sheet_write.append(masked_row)
# 			else:
# 				sheet_write.append(row_values)
# 			current_count += 1
# 			if current_count % 500 == 0:
# 				print('完成了{}%'.format(current_count / row_count * 100))
# 	save_pickle('mapping.pkl', hash_bytes)
# 	wb_write.save('17.18脱敏数据.xlsx')



if __name__ == '__main__':

	# a = 'asdasd'
	# d = {}
	# d[hash(a)] = a
	# print(d)
	#
	#
	#
	# hash_bytes = {}
	# data_dict = read_xlsx('../17、18年统计数据格式模板.xlsx')
	# data_dict_enc = {}
	# for year, data_rows in data_dict.items():
	# 	year_, data_rows_, hash_bytes_ = do_encrypt(append_zero('123'), data_rows, year)
	# 	hash_bytes.update(hash_bytes_)
	# 	print(year, hash_bytes)
	# 	data_dict_enc[year_] = data_rows_
	# 	for data_row in data_rows_:
	# 		print(data_row)
	# save_xlsx('../加密数据.xlsx', data_dict_enc)
	# save_pickle('mapping.pkl', hash_bytes)
	#
	# data_dict_enc = read_xlsx('../加密数据.xlsx')
	# hash_bytes = load_pickle('mapping.pkl')
	# for year, data_rows in data_dict_enc.items():
	# 	print(do_decrypt(append_zero('123'), data_rows, year, hash_bytes))
	#
	# a = {}
	# a['1'] = 10
	# b = {}
	# b['1'] = 100
	# a.update(b)
	# print(a)


	wb_read = load_workbook('17.18_bigdata.xlsx', read_only=True)
	wb_write = Workbook(write_only=True)
	hash_bytes = load_pickle('mapping.pkl')
	# get the count of all data rows
	row_count = 0
	current_count = 1
	for sheetname in wb_read.sheetnames:
		sheet_read = wb_read[sheetname]
		row_count += sheet_read.max_row
	# read data and do masking, and then save the masked rows
	for sheetname in wb_read.sheetnames:
		print('processing sheet {}:'.format(sheetname))
		sheet_read = wb_read[sheetname]
		# sheet_row_count = sheet_read.max_row
		# print(sheet_row_count)
		sheet_write = wb_write.create_sheet(title=sheetname)
		rows_read = sheet_read.rows
		for row in rows_read:
			time_from = time.time()
			row_values = []
			for cell in row:
				row_values.append(cell.value)
			# do masking
			if current_count > 1:
				# masked_row, hash_bytes_added = mask_row('1234567890123456', sheetname, row_values)
				# hash_bytes.update(hash_bytes_added)
				for i in range(15000):
					sheet_write.append(row_values)
			else:
				sheet_write.append(row_values)
			current_count += 1
			time_to = time.time()
			print('完成了{}%, 用时{}s'.format(current_count / row_count * 100, time_to - time_from))
	wb_write.save('17.18_bigdata_1.xlsx')