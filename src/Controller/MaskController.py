'''
--------------------------------------------------------
@File    :   MaskerQMainWindow.py    
@Contact :   1183862787@qq.com
@License :   (C)Copyright 2017-2018, CS, WHU

@Modify Time : 2019/5/19 9:21     
@Author      : Liu Wang    
@Version     : 1.0   
@Desciption  : None
--------------------------------------------------------  
''' 

from Viewer.main_window import Ui_MainWindow
from PyQt5.QtWidgets import QMainWindow, QFileDialog, QMessageBox, QWidget
from PyQt5.QtCore import pyqtSlot
from openpyxl import load_workbook, Workbook
from src.utils import append_zero, save_pickle, load_pickle, mask_row, demask_row

class MaskController():

	def __init__(self):
		self.mainWindow = QMainWindow()
		self.ui = Ui_MainWindow()
		self.ui.setupUi(self.mainWindow)
		self.ui.pushButton_export.clicked.connect(lambda: self._do_masking())
		self.ui.pushButton_browse.clicked.connect(lambda: self._browse_file())
		self.ui.pushButton_import.clicked.connect(lambda: self._do_demasking())

	@pyqtSlot()
	def _do_masking(self):
		try:
			# get key and dir's abs_path
			key_str = self._get_key()
			file_path = self._get_path()
			# print(key_str, file_path)
			wb_read = load_workbook(file_path, read_only=True)
			wb_write = Workbook(write_only=True)
			hash_bytes = load_pickle('mapping.pkl')
			# get the count of all data rows
			row_count = 0
			current_count = 1
			for sheetname in wb_read.sheetnames:
				sheet_read = wb_read[sheetname]
				row_count += sheet_read.max_row
			# read data and do masking, and then save the masked rows
			for sheetname in wb_read.sheetnames:
				print('processing sheet {}:'.format(sheetname))
				sheet_read = wb_read[sheetname]
				# sheet_row_count = sheet_read.max_row
				# print(sheet_row_count)
				sheet_write = wb_write.create_sheet(title=sheetname)
				rows_read = sheet_read.rows
				for row in rows_read:
					row_values = []
					for cell in row:
						row_values.append(cell.value)
					# do masking
					if current_count > 1:
						masked_row, hash_bytes_added = mask_row(key_str, sheetname, row_values)
						hash_bytes.update(hash_bytes_added)
						sheet_write.append(masked_row)
					else:
						sheet_write.append(row_values)
					current_count += 1
					if current_count % 100 == 0 or current_count == row_count:
						self._set_processBar(current_count / row_count * 100)
						print('完成了{}%'.format(current_count / row_count * 100))
			save_pickle('mapping.pkl', hash_bytes)

			QMessageBox.information(QWidget(), "Information", "数据脱敏成功，点击确认后请保存文件")
			write_path = QFileDialog.getSaveFileName(caption="保存为.xlsx文档", directory="./")[0]
			# write_path = './加密数据.xlsx'
			# print(write_path)
			wb_write.save(write_path)
			QMessageBox.information(QWidget(), "Information", "保存完成")
		except Exception as e:
			QMessageBox.warning(QWidget(), "warning", str(e))
			print(e)

	@pyqtSlot()
	def _do_demasking(self):
		try:
			# print('do import')
			# get key and dir's abs_path
			key_str = self._get_key()
			file_path = self._get_path()
			# print(key_str, file_path)

			QMessageBox.information(QWidget(), "Information", "解密成功，点击确认后请保存文件")
			write_path = QFileDialog.getSaveFileName(caption="保存为.xlsx文档", directory="./")[0]
		except Exception as e:
			QMessageBox.warning(QWidget(), "warning", str(e))
			# print(e)


	@pyqtSlot()
	def _browse_dir(self):
		selected_path = QFileDialog.getExistingDirectory(caption="浏览", directory="./")
		self.ui.lineEdit_dir.setText(selected_path)

	@pyqtSlot()
	def _browse_file(self):
		selected_path = QFileDialog.getOpenFileName(caption="浏览", directory="./")
		self.ui.lineEdit_dir.setText(selected_path[0])

	def _set_processBar(self, value):
		self.ui.progressBar.setValue(int(value))

	def _get_path(self):
		dir_path = self.ui.lineEdit_dir.text()
		if len(dir_path) <= 0:
			raise Exception('请选择要加密导出的文件!')
		else:
			return dir_path

	def _get_key(self):
		return append_zero(self.ui.lineEdit_key.text())